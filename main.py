"""
Script principal para probar el sistema de selección de transportistas

Demuestra las diferentes funcionalidades:
- Selección del mejor transportista
- Comparación de todas las opciones
- Visualización detallada de cotizaciones
"""

import sys
from pathlib import Path
from decimal import Decimal
from datetime import datetime

# Añadir el directorio raíz al path
sys.path.insert(0, str(Path(__file__).parent))

from database import get_db_manager
from services import TransportistaSelector
from models import Pedido, Transportista, ServicioTransportista, Tarifa, TipoEntrega, MetodoCalculo

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False


def imprimir_separador(caracter="=", longitud=80):
    """Imprime una línea separadora"""
    print(caracter * longitud)


def imprimir_cotizacion(cotizacion, posicion=None):
    """Imprime una cotización de forma bonita"""
    prefijo = f"{posicion}. " if posicion else "• "
    print(f"{prefijo}{cotizacion.transportista_nombre} - {cotizacion.tipo_entrega.replace('_', ' ').title()}")
    print(f"   Método: {cotizacion.metodo_calculo.upper()}")
    print(f"   Cantidad: {cotizacion.cantidad_calculada:.2f}")
    print(f"   Precio: {cotizacion.precio_total:.2f}€")
    print(f"   Desglose: {cotizacion.detalles}")
    print(f"   Provincia tarifa: {cotizacion.provincia}")
    print()


def mostrar_mejor_transportista(pedido_id: int, session):
    """Muestra el mejor transportista para un pedido"""
    selector = TransportistaSelector(session)
    
    # Obtener información del pedido
    pedido = session.query(Pedido).filter(Pedido.id == pedido_id).first()
    
    imprimir_separador()
    print(f"🎯 MEJOR TRANSPORTISTA - Pedido {pedido.numero_pedido}")
    imprimir_separador()
    print(f"Provincia: {pedido.provincia_entrega}")
    print(f"Tipo de entrega: {pedido.tipo_entrega.value.replace('_', ' ').title()}")
    print()
    
    # Calcular totales
    totales = selector.calcular_totales_pedido(pedido)
    print(f"📦 TOTALES DEL PEDIDO:")
    print(f"   Peso total: {totales['peso_total']:.2f} kg")
    print(f"   Volumen total: {totales['volumen_total']:.4f} m³")
    print(f"   Palets estimados: {totales['palets_total']:.2f}")
    print()
    
    print("📋 PRODUCTOS:")
    for pp in pedido.productos:
        print(f"   • {pp.producto.nombre} x{pp.cantidad}")
        print(f"     ({pp.producto.peso_kg}kg, {pp.producto.volumen_m3}m³ c/u)")
    print()
    
    # Seleccionar mejor transportista
    mejor = selector.seleccionar_mejor_transportista(pedido_id)
    
    if mejor:
        print("✅ MEJOR OPCIÓN:")
        imprimir_cotizacion(mejor)
    else:
        print("❌ No hay transportistas disponibles para este pedido")
    
    imprimir_separador()
    print()


def comparar_transportistas(pedido_id: int, session):
    """Compara todos los transportistas disponibles para un pedido"""
    selector = TransportistaSelector(session)
    comparacion = selector.comparar_transportistas(pedido_id)
    
    pedido_info = comparacion['pedido']
    mejor = comparacion['mejor_opcion']
    todas = comparacion['todas_cotizaciones']
    ahorro = comparacion['ahorro_mejor_vs_peor']
    
    imprimir_separador()
    print(f"📊 COMPARACIÓN DE TRANSPORTISTAS - Pedido {pedido_info['numero']}")
    imprimir_separador()
    print(f"Provincia: {pedido_info['provincia']} | Tipo de entrega: {pedido_info['tipo_entrega'].replace('_', ' ').title()} | Peso: {pedido_info['peso_total_kg']:.2f} kg | Volumen: {pedido_info['volumen_total_m3']:.4f} m³ | Palets: {pedido_info['palets_total']:.2f}")
    print()
    
    if not todas:
        print("❌ No hay transportistas disponibles para este pedido")
        imprimir_separador()
        return
    
    print(f"🏆 RANKING DE OPCIONES ({len(todas)} disponibles):")
    print()
    
    # Encabezado de la tabla
    print(f"{'#':>3} {'Transportista':<20} {'Tipo Entrega':<25} {'Método':<8} {'Cantidad':>10} {'Precio':>10} {'Provincia Tarifa':<17}")
    print("=" * 120)
    
    # Datos
    for i, cotizacion in enumerate(todas, 1):
        posicion = f"{i}."
        tipo_entrega = cotizacion.tipo_entrega.replace('_', ' ').title()
        metodo = cotizacion.metodo_calculo.upper()
        cantidad = cotizacion.cantidad_calculada
        precio = cotizacion.precio_total
        provincia = cotizacion.provincia
        
        # Formatear cantidad según el método
        if metodo == "VOLUMEN":
            cantidad_str = f"{cantidad:.2f} m³"
        elif metodo == "PESO":
            cantidad_str = f"{cantidad:.2f} kg"
        else:  # PALETS
            cantidad_str = f"{cantidad:.2f} pal"
        
        print(f"{posicion:>3} {cotizacion.transportista_nombre:<20} {tipo_entrega:<25} {metodo:<8} {cantidad_str:>10} {precio:>9.2f}€ {provincia:<17}")
    
    print()
    if len(todas) > 1:
        print(f"💰 AHORRO: {ahorro:.2f}€ eligiendo la mejor opción ({(ahorro / float(todas[-1].precio_total)) * 100:.1f}% más económico que la opción más cara)")
    
    imprimir_separador()
    print()


def exportar_tarifas_excel(session):
    """Exporta todas las tarifas a un archivo Excel"""
    if not EXCEL_DISPONIBLE:
        print("\n❌ ERROR: La librería openpyxl no está instalada.")
        print("Instala con: pip install openpyxl")
        return
    
    try:
        # Crear libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Tarifas"
        
        # Estilo del encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            "ID", "Transportista", "Servicio", "Tipo Entrega", "Método Cálculo",
            "Provincia", "Rango Min", "Rango Max", "Precio Fijo"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Obtener todas las tarifas con sus relaciones
        tarifas = session.query(Tarifa).join(
            ServicioTransportista
        ).join(
            Transportista
        ).order_by(
            Transportista.nombre,
            ServicioTransportista.tipo_entrega,
            Tarifa.provincia,
            Tarifa.rango_min
        ).all()
        
        # Escribir datos
        for row, tarifa in enumerate(tarifas, 2):
            servicio = tarifa.servicio
            transportista = servicio.transportista
            
            # Generar nombre descriptivo del servicio
            servicio_nombre = f"{servicio.tipo_entrega.value.replace('_', ' ').title()} ({servicio.metodo_calculo.value})"
            
            ws.cell(row=row, column=1, value=tarifa.id)
            ws.cell(row=row, column=2, value=transportista.nombre)
            ws.cell(row=row, column=3, value=servicio_nombre)
            ws.cell(row=row, column=4, value=servicio.tipo_entrega.value)
            ws.cell(row=row, column=5, value=servicio.metodo_calculo.value)
            ws.cell(row=row, column=6, value=tarifa.provincia)
            ws.cell(row=row, column=7, value=float(tarifa.rango_min))
            ws.cell(row=row, column=8, value=float(tarifa.rango_max) if tarifa.rango_max else "")
            ws.cell(row=row, column=9, value=float(tarifa.precio_fijo))
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"tarifas_export_{timestamp}.xlsx"
        wb.save(filename)
        
        print(f"\n✅ Tarifas exportadas correctamente")
        print(f"📁 Archivo: {filename}")
        print(f"📊 Total de tarifas: {len(tarifas)}")
        
    except Exception as e:
        print(f"\n❌ Error al exportar tarifas: {e}")


def importar_tarifas_excel(session):
    """Importa tarifas desde un archivo Excel"""
    if not EXCEL_DISPONIBLE:
        print("\n❌ ERROR: La librería openpyxl no está instalada.")
        print("Instala con: pip install openpyxl")
        return
    
    filename = input("\nNombre del archivo Excel (ej: tarifas_export_20241216.xlsx): ").strip()
    
    if not Path(filename).exists():
        print(f"\n❌ Error: El archivo '{filename}' no existe.")
        return
    
    try:
        # Cargar el archivo
        wb = load_workbook(filename)
        ws = wb.active
        
        # Verificar encabezados
        headers_esperados = [
            "ID", "Transportista", "Servicio", "Tipo Entrega", "Método Cálculo",
            "Provincia", "Rango Min", "Rango Max", "Precio Fijo"
        ]
        
        headers_archivo = [cell.value for cell in ws[1]]
        if headers_archivo != headers_esperados:
            print("\n❌ Error: El formato del archivo no es correcto.")
            print(f"Encabezados esperados: {headers_esperados}")
            print(f"Encabezados encontrados: {headers_archivo}")
            return
        
        # Procesar filas
        tarifas_nuevas = 0
        tarifas_actualizadas = 0
        errores = []
        
        for row_idx in range(2, ws.max_row + 1):
            try:
                tarifa_id = ws.cell(row=row_idx, column=1).value
                transportista_nombre = ws.cell(row=row_idx, column=2).value
                servicio_nombre = ws.cell(row=row_idx, column=3).value
                tipo_entrega_str = ws.cell(row=row_idx, column=4).value
                metodo_calculo_str = ws.cell(row=row_idx, column=5).value
                provincia = ws.cell(row=row_idx, column=6).value
                rango_min = ws.cell(row=row_idx, column=7).value
                rango_max = ws.cell(row=row_idx, column=8).value
                precio_fijo = ws.cell(row=row_idx, column=9).value
                
                # Validar datos obligatorios
                if not all([transportista_nombre, servicio_nombre, tipo_entrega_str, 
                           metodo_calculo_str, provincia, rango_min is not None, precio_fijo]):
                    errores.append(f"Fila {row_idx}: Faltan datos obligatorios")
                    continue
                
                # Buscar transportista
                transportista = session.query(Transportista).filter(
                    Transportista.nombre == transportista_nombre
                ).first()
                
                if not transportista:
                    errores.append(f"Fila {row_idx}: Transportista '{transportista_nombre}' no encontrado")
                    continue
                
                # Convertir enums
                try:
                    tipo_entrega = TipoEntrega(tipo_entrega_str)
                    metodo_calculo = MetodoCalculo(metodo_calculo_str)
                except ValueError:
                    errores.append(f"Fila {row_idx}: Tipo de entrega o método de cálculo inválido")
                    continue
                
                # Buscar servicio
                servicio = session.query(ServicioTransportista).filter(
                    ServicioTransportista.transportista_id == transportista.id,
                    ServicioTransportista.tipo_entrega == tipo_entrega,
                    ServicioTransportista.metodo_calculo == metodo_calculo
                ).first()
                
                if not servicio:
                    errores.append(f"Fila {row_idx}: Servicio '{servicio_nombre}' no encontrado para '{transportista_nombre}'")
                    continue
                
                # Convertir rango_max
                rango_max_decimal = Decimal(str(rango_max)) if rango_max else None
                
                # Verificar si la tarifa existe (por ID o por combinación única)
                tarifa_existente = None
                if tarifa_id:
                    tarifa_existente = session.query(Tarifa).filter(Tarifa.id == tarifa_id).first()
                
                if not tarifa_existente:
                    # Buscar por combinación única
                    tarifa_existente = session.query(Tarifa).filter(
                        Tarifa.servicio_id == servicio.id,
                        Tarifa.provincia == provincia,
                        Tarifa.rango_min == Decimal(str(rango_min)),
                        Tarifa.rango_max == rango_max_decimal if rango_max_decimal else Tarifa.rango_max.is_(None)
                    ).first()
                
                if tarifa_existente:
                    # Actualizar tarifa existente
                    tarifa_existente.precio_fijo = Decimal(str(precio_fijo))
                    tarifas_actualizadas += 1
                else:
                    # Crear nueva tarifa
                    nueva_tarifa = Tarifa(
                        servicio_id=servicio.id,
                        provincia=provincia,
                        rango_min=Decimal(str(rango_min)),
                        rango_max=rango_max_decimal,
                        precio_fijo=Decimal(str(precio_fijo))
                    )
                    session.add(nueva_tarifa)
                    tarifas_nuevas += 1
                    
            except Exception as e:
                errores.append(f"Fila {row_idx}: {str(e)}")
        
        # Confirmar cambios
        if tarifas_nuevas > 0 or tarifas_actualizadas > 0:
            confirmacion = input(f"\n¿Confirmar importación? ({tarifas_nuevas} nuevas, {tarifas_actualizadas} actualizadas) (S/n): ").strip().lower()
            if confirmacion in ['s', 'si', 'sí', '']:
                session.commit()
                print(f"\n✅ Importación completada")
                print(f"   📝 Tarifas nuevas: {tarifas_nuevas}")
                print(f"   🔄 Tarifas actualizadas: {tarifas_actualizadas}")
            else:
                session.rollback()
                print("\n❌ Importación cancelada")
        else:
            print("\n⚠️ No se encontraron cambios para aplicar")
        
        if errores:
            print(f"\n⚠️ Se encontraron {len(errores)} errores:")
            for error in errores[:10]:  # Mostrar solo los primeros 10
                print(f"   • {error}")
            if len(errores) > 10:
                print(f"   ... y {len(errores) - 10} errores más")
                
    except Exception as e:
        print(f"\n❌ Error al importar tarifas: {e}")
        session.rollback()


def listar_tarifas(session):
    """Lista todas las tarifas en formato tabular"""
    print("\n💰 LISTADO DE TARIFAS")
    print("=" * 140)
    
    # Obtener todas las tarifas con sus relaciones
    tarifas = session.query(Tarifa).join(
        ServicioTransportista
    ).join(
        Transportista
    ).filter(
        Transportista.activo == True,
        ServicioTransportista.activo == True
    ).order_by(
        Transportista.nombre,
        ServicioTransportista.tipo_entrega,
        Tarifa.provincia,
        Tarifa.rango_min
    ).all()
    
    if not tarifas:
        print("❌ No hay tarifas disponibles")
        return
    
    # Encabezado de la tabla
    print(f"\n{'Transportista':<15} {'Tipo Entrega':<22} {'Método':<8} {'Provincia':<12} {'Rango Min':>10} {'Rango Max':>10} {'Precio':>10}")
    print("=" * 140)
    
    # Datos agrupados por transportista
    transportista_actual = None
    tipo_entrega_actual = None
    
    for tarifa in tarifas:
        servicio = tarifa.servicio
        transportista = servicio.transportista
        
        # Determinar unidad
        unidad = {
            MetodoCalculo.PESO: 'kg',
            MetodoCalculo.VOLUMEN: 'm³',
            MetodoCalculo.PALETS: 'pal'
        }[servicio.metodo_calculo]
        
        # Formatear valores
        transportista_nombre = transportista.nombre if transportista.nombre != transportista_actual else ""
        tipo_entrega_str = servicio.tipo_entrega.value.replace('_', ' ').title()
        
        if transportista.nombre != transportista_actual or servicio.tipo_entrega.value != tipo_entrega_actual:
            tipo_display = tipo_entrega_str
        else:
            tipo_display = ""
        
        metodo_str = servicio.metodo_calculo.value.upper()
        rango_min_str = f"{tarifa.rango_min:.2f} {unidad}"
        rango_max_str = f"{tarifa.rango_max:.2f} {unidad}" if tarifa.rango_max else "∞"
        precio_str = f"{tarifa.precio_fijo:.2f}€"
        
        # Imprimir fila
        print(f"{transportista_nombre:<15} {tipo_display:<22} {metodo_str:<8} {tarifa.provincia:<12} {rango_min_str:>10} {rango_max_str:>10} {precio_str:>10}")
        
        # Actualizar variables de control
        transportista_actual = transportista.nombre
        tipo_entrega_actual = servicio.tipo_entrega.value
        
        # Línea separadora entre transportistas
        if transportista_nombre:
            siguiente_idx = tarifas.index(tarifa) + 1
            if siguiente_idx < len(tarifas):
                siguiente_transportista = tarifas[siguiente_idx].servicio.transportista.nombre
                if siguiente_transportista != transportista_actual:
                    print("-" * 140)
    
    print("=" * 140)
    print(f"\n📊 Total de tarifas: {len(tarifas)}")
    print()


def menu_principal():
    """Menú interactivo principal"""
    db_manager = get_db_manager()
    
    while True:
        print("\n" + "=" * 60)
        print("🚚 SISTEMA DE SELECCIÓN DE TRANSPORTISTAS")
        print("=" * 60)
        print("\nOpciones:")
        print("1. Ver mejor transportista para cada pedido")
        print("2. Comparar transportistas para un pedido específico")
        print("3. Comparar transportistas para TODOS los pedidos")
        print("4. Listar todos los pedidos")
        print("5. Listar tarifas de todos los transportistas")
        print("6. Exportar tarifas a Excel")
        print("7. Importar tarifas desde Excel")
        print("8. Salir")
        print()
        
        opcion = input("Selecciona una opción (1-8): ").strip()
        
        if opcion == '1':
            with db_manager.get_session() as session:
                pedidos = session.query(Pedido).all()
                for pedido in pedidos:
                    mostrar_mejor_transportista(pedido.id, session)
                    input("Presiona ENTER para continuar...")
        
        elif opcion == '2':
            with db_manager.get_session() as session:
                pedidos = session.query(Pedido).all()
                print("\nPedidos disponibles:")
                for i, p in enumerate(pedidos, 1):
                    print(f"{i}. {p.numero_pedido} - {p.provincia_entrega} - {p.tipo_entrega.value}")
                
                try:
                    seleccion = int(input("\nSelecciona un pedido (número): ")) - 1
                    if 0 <= seleccion < len(pedidos):
                        comparar_transportistas(pedidos[seleccion].id, session)
                        input("Presiona ENTER para continuar...")
                    else:
                        print("❌ Selección inválida")
                except ValueError:
                    print("❌ Entrada inválida")
        
        elif opcion == '3':
            with db_manager.get_session() as session:
                pedidos = session.query(Pedido).all()
                for pedido in pedidos:
                    comparar_transportistas(pedido.id, session)
                    input("Presiona ENTER para continuar...")
        
        elif opcion == '4':
            with db_manager.get_session() as session:
                selector = TransportistaSelector(session)
                pedidos = session.query(Pedido).all()
                print("\n📦 LISTADO DE PEDIDOS:")
                print()
                # Encabezado de la tabla
                print(f"{'Pedido':<15} {'Provincia':<15} {'Tipo Entrega':<25} {'Prods':>5} {'Peso (kg)':>10} {'Volumen (m³)':>13} {'Palets':>8}")
                print("=" * 120)
                # Datos
                for pedido in pedidos:
                    totales = selector.calcular_totales_pedido(pedido)
                    tipo_entrega = pedido.tipo_entrega.value.replace('_', ' ').title()
                    print(f"{pedido.numero_pedido:<15} {pedido.provincia_entrega:<15} {tipo_entrega:<25} {len(pedido.productos):>5} {totales['peso_total']:>10.2f} {totales['volumen_total']:>13.4f} {totales['palets_total']:>8.2f}")
                print()
                input("Presiona ENTER para continuar...")
        
        elif opcion == '5':
            with db_manager.get_session() as session:
                listar_tarifas(session)
                input("Presiona ENTER para continuar...")
        
        elif opcion == '6':
            with db_manager.get_session() as session:
                exportar_tarifas_excel(session)
                input("Presiona ENTER para continuar...")
        
        elif opcion == '7':
            with db_manager.get_session() as session:
                importar_tarifas_excel(session)
                input("Presiona ENTER para continuar...")
        
        elif opcion == '8':
            print("\n👋 ¡Hasta luego!")
            break
        
        else:
            print("❌ Opción inválida")


def demo_rapido():
    """Demostración rápida del sistema"""
    print("\n" + "=" * 60)
    print("🚀 DEMOSTRACIÓN RÁPIDA DEL SISTEMA")
    print("=" * 60)
    print()
    
    db_manager = get_db_manager()
    
    with db_manager.get_session() as session:
        pedidos = session.query(Pedido).all()
        
        print(f"Se analizarán {len(pedidos)} pedidos de ejemplo...\n")
        
        for pedido in pedidos:
            comparar_transportistas(pedido.id, session)
            print()


if __name__ == "__main__":
    # Verificar si la base de datos existe
    db_path = Path(__file__).parent / "transportistas.db"
    
    if not db_path.exists():
        print("\n⚠️  La base de datos no existe.")
        print("Por favor, ejecuta primero: python init_db.py\n")
        sys.exit(1)
    
    # Preguntar modo de ejecución
    print("\n¿Cómo deseas ejecutar el programa?")
    print("1. Demostración rápida (muestra todos los pedidos)")
    print("2. Menú interactivo")
    
    modo = input("\nSelecciona (1 o 2): ").strip()
    
    if modo == '1':
        demo_rapido()
    else:
        menu_principal()
